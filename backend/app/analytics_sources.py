"""Analytics data sources: one adapter interface over datasets and SQL tables.

A *source* is a registered pointer to tabular data. Internal datasets are
implicit sources (always listed, nothing stored); SQL sources pair an
encrypted credential (the same ones the Postgres connector uses) with a
table name. Every kind answers the same three questions — schema, rows,
aggregate — which is exactly what the Analytics UI consumes.
"""

from __future__ import annotations

import json
import logging
import re
import sqlite3
import uuid
from datetime import UTC, datetime
from typing import Any

from app.config import get_settings
from app.credentials import get_credential_store
from genxai.core.datasets import ALLOWED_METRICS, aggregate_rows, get_dataset_store
from genxai.core.files import get_file_store

logger = logging.getLogger(__name__)

_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]{0,62}$")
MAX_SQL_ROWS = 500
MAX_GROUPS = 50

_METRIC_SQL = {
    "count": "COUNT(*)",
    "sum": "SUM({field})",
    "avg": "AVG({field})",
    "min": "MIN({field})",
    "max": "MAX({field})",
}


def _validate_identifier(name: str, kind: str) -> str:
    if not _IDENTIFIER_RE.match(name or ""):
        raise ValueError(f"Invalid {kind} name {name!r}")
    return name


# ------------------------------------------------------------------ registry


class SourceRegistry:
    """Registered (non-implicit) sources, stored in the datasets database."""

    def __init__(self) -> None:
        self.db_path = get_settings().data_dir / "datasets.db"
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS sources (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    kind TEXT NOT NULL,
                    config TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
                """
            )

    def list(self) -> list[dict[str, Any]]:
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute(
                "SELECT id, name, kind, config, created_at FROM sources "
                "ORDER BY created_at DESC"
            )
            return [
                {
                    "id": row[0],
                    "name": row[1],
                    "kind": row[2],
                    "config": json.loads(row[3]),
                    "created_at": row[4],
                }
                for row in cursor.fetchall()
            ]

    def get(self, source_id: str) -> dict[str, Any] | None:
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(
                "SELECT id, name, kind, config, created_at FROM sources WHERE id = ?",
                (source_id,),
            ).fetchone()
        if row is None:
            return None
        return {
            "id": row[0],
            "name": row[1],
            "kind": row[2],
            "config": json.loads(row[3]),
            "created_at": row[4],
        }

    def create(self, name: str, kind: str, config: dict[str, Any]) -> dict[str, Any]:
        source = {
            "id": uuid.uuid4().hex,
            "name": name,
            "kind": kind,
            "config": config,
            "created_at": datetime.now(UTC).isoformat(),
        }
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                "INSERT INTO sources (id, name, kind, config, created_at) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    source["id"],
                    name,
                    kind,
                    json.dumps(config),
                    source["created_at"],
                ),
            )
        return source

    def delete(self, source_id: str) -> bool:
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute("DELETE FROM sources WHERE id = ?", (source_id,))
        return cursor.rowcount > 0


_registry: SourceRegistry | None = None


def get_source_registry() -> SourceRegistry:
    global _registry
    if _registry is None:
        _registry = SourceRegistry()
    return _registry


def reset_source_registry() -> None:
    global _registry
    _registry = None


# ------------------------------------------------------------------ adapters


class DatasetAdapter:
    """Internal dataset store — wraps the existing DatasetStore calls."""

    def __init__(self, dataset: str) -> None:
        self.dataset = dataset

    def schema(self) -> list[dict[str, str]]:
        sample = get_dataset_store().rows(self.dataset, limit=50)["rows"]
        return _infer_schema(sample)

    def rows(self, limit: int, offset: int) -> dict[str, Any]:
        return get_dataset_store().rows(self.dataset, limit=limit, offset=offset)

    def aggregate(
        self, metric: str, field: str | None, group_by: str | None
    ) -> list[dict[str, Any]]:
        return get_dataset_store().aggregate(
            self.dataset, metric=metric, field=field, group_by=group_by
        )


def validate_readonly_sql(sql: str) -> str:
    """Accept only single SELECT/WITH statements for analytics sources."""
    cleaned = (sql or "").strip().rstrip(";")
    first_word = cleaned.split(None, 1)
    if not first_word or first_word[0].lower() not in ("select", "with"):
        raise ValueError("Custom SQL must be a SELECT/WITH query")
    if ";" in cleaned:
        raise ValueError("Custom SQL must be a single statement")
    return cleaned


class SQLAdapter:
    """A table (or read-only custom query) behind a stored credential.

    Aggregation is pushed down as SQL; custom queries are wrapped as a
    subquery so paging and GROUP BY run in the database either way.
    """

    def __init__(
        self, credential: str, table: str | None = None, sql: str | None = None
    ) -> None:
        entry = get_credential_store().get(credential)
        if entry is None:
            raise LookupError(f"Credential '{credential}' not found")
        connection_string = entry.config.get("connection_string")
        if not connection_string:
            raise LookupError(
                f"Credential '{credential}' has no connection_string"
            )
        self.connection_string = str(connection_string)
        if bool(table) == bool(sql):
            raise ValueError("Provide exactly one of table / sql")
        self.table = _validate_identifier(table, "table") if table else None
        self.sql = validate_readonly_sql(sql) if sql else None

    @property
    def _relation(self) -> str:
        """What goes after FROM: the table name or the wrapped query."""
        return self.table if self.table else f"({self.sql}) AS _q"

    def _engine(self) -> Any:
        from sqlalchemy import create_engine

        return create_engine(self.connection_string, pool_pre_ping=True)

    def schema(self) -> list[dict[str, str]]:
        from sqlalchemy import inspect, text

        engine = self._engine()
        try:
            if self.table:
                columns = inspect(engine).get_columns(self.table)
                return [
                    {"name": column["name"], "type": str(column["type"]).lower()}
                    for column in columns
                ]
            with engine.connect() as conn:
                result = conn.execute(
                    text(f"SELECT * FROM {self._relation} LIMIT 25")  # noqa: S608
                )
                columns_names = list(result.keys())
                sample = [
                    {column: value for column, value in zip(columns_names, row)}
                    for row in result.fetchall()
                ]
            inferred = _infer_schema(sample)
            known = {entry["name"] for entry in inferred}
            inferred.extend(
                {"name": name, "type": "string"}
                for name in columns_names
                if name not in known
            )
            return inferred
        finally:
            engine.dispose()

    def rows(self, limit: int, offset: int) -> dict[str, Any]:
        from sqlalchemy import text

        engine = self._engine()
        try:
            with engine.connect() as conn:
                total = conn.execute(
                    text(f"SELECT COUNT(*) FROM {self._relation}")  # noqa: S608
                ).scalar()
                result = conn.execute(
                    text(
                        f"SELECT * FROM {self._relation} LIMIT :limit OFFSET :offset"  # noqa: S608
                    ),
                    {"limit": min(limit, MAX_SQL_ROWS), "offset": offset},
                )
                columns = list(result.keys())
                rows = [
                    {column: value for column, value in zip(columns, row)}
                    for row in result.fetchall()
                ]
        finally:
            engine.dispose()
        return {"rows": rows, "total": int(total or 0)}

    def aggregate(
        self, metric: str, field: str | None, group_by: str | None
    ) -> list[dict[str, Any]]:
        from sqlalchemy import text

        if metric not in ALLOWED_METRICS:
            raise ValueError(f"metric must be one of {ALLOWED_METRICS}")
        if metric != "count":
            if not field:
                raise ValueError(f"metric '{metric}' requires a field")
            _validate_identifier(field, "column")
        agg_sql = _METRIC_SQL[metric].format(field=field)

        engine = self._engine()
        try:
            with engine.connect() as conn:
                if group_by:
                    _validate_identifier(group_by, "column")
                    statement = (
                        f"SELECT {group_by} AS grp, {agg_sql} AS value, "  # noqa: S608
                        f"COUNT(*) AS rows_in_group FROM {self._relation} "
                        f"GROUP BY {group_by} ORDER BY value DESC LIMIT {MAX_GROUPS}"
                    )
                    result = conn.execute(text(statement))
                    return [
                        {
                            "group": "—" if grp is None else str(grp),
                            "value": float(value or 0),
                            "rows": int(count),
                        }
                        for grp, value, count in result.fetchall()
                    ]
                statement = (
                    f"SELECT {agg_sql} AS value, COUNT(*) AS n FROM {self._relation}"  # noqa: S608
                )
                value, count = conn.execute(text(statement)).fetchone()
                return [
                    {
                        "group": "all",
                        "value": float(value or 0),
                        "rows": int(count),
                    }
                ]
        finally:
            engine.dispose()


MAX_FILE_ROWS = 50_000
_file_cache: dict[str, list[dict[str, Any]]] = {}
_FILE_CACHE_MAX = 8


def _coerce(value: str) -> Any:
    """CSV cells are all strings; give numbers their type back."""
    text = value.strip()
    if text == "":
        return None
    try:
        return int(text)
    except ValueError:
        try:
            return float(text)
        except ValueError:
            return value


def parse_file_rows(
    file_id: str, format_: str, sheet: str | None = None
) -> list[dict[str, Any]]:
    """Parse an xlsx/csv file from the file store into rows (cached by content)."""
    cache_key = f"{file_id}:{format_}:{sheet or ''}"
    if cache_key in _file_cache:
        return _file_cache[cache_key]

    data = get_file_store().read_bytes(file_id)
    rows: list[dict[str, Any]] = []
    if format_ == "csv":
        import csv
        import io

        reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig", errors="replace")))
        for record in reader:
            if len(rows) >= MAX_FILE_ROWS:
                break
            rows.append(
                {key: _coerce(value) if isinstance(value, str) else value
                 for key, value in record.items() if key is not None}
            )
    elif format_ == "xlsx":
        import io

        import openpyxl

        workbook = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
        try:
            names = workbook.sheetnames
            if sheet and sheet not in names:
                raise ValueError(f"Sheet {sheet!r} not found — workbook has {names}")
            ws = workbook[sheet] if sheet else workbook[names[0]]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is not None:
                columns = [
                    str(cell) if cell is not None else f"col_{index + 1}"
                    for index, cell in enumerate(header)
                ]
                for values in rows_iter:
                    if len(rows) >= MAX_FILE_ROWS:
                        break
                    if all(value is None for value in values):
                        continue
                    rows.append(
                        {
                            columns[index] if index < len(columns) else f"col_{index + 1}":
                                str(value) if hasattr(value, "isoformat") else value
                            for index, value in enumerate(values)
                        }
                    )
        finally:
            workbook.close()
    else:
        raise ValueError(f"Unsupported file format '{format_}' (xlsx or csv)")

    if len(_file_cache) >= _FILE_CACHE_MAX:
        _file_cache.pop(next(iter(_file_cache)))
    _file_cache[cache_key] = rows
    return rows


def list_workbook_sheets(file_id: str) -> list[str]:
    """Sheet names of an .xlsx in the file store (for the upload dialog)."""
    import io

    import openpyxl

    workbook = openpyxl.load_workbook(
        io.BytesIO(get_file_store().read_bytes(file_id)), read_only=True
    )
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


class FileAdapter:
    """An uploaded xlsx/csv file — parsed once, served from memory."""

    def __init__(self, file_id: str, format_: str, sheet: str | None = None) -> None:
        self.rows_data = parse_file_rows(file_id, format_, sheet)

    def schema(self) -> list[dict[str, str]]:
        return _infer_schema(self.rows_data[:100])

    def rows(self, limit: int, offset: int) -> dict[str, Any]:
        return {
            "rows": self.rows_data[offset : offset + limit],
            "total": len(self.rows_data),
        }

    def aggregate(
        self, metric: str, field: str | None, group_by: str | None
    ) -> list[dict[str, Any]]:
        return aggregate_rows(
            self.rows_data, metric=metric, field=field, group_by=group_by
        )


def _infer_schema(sample: list[dict[str, Any]]) -> list[dict[str, str]]:
    types: dict[str, str] = {}
    for row in sample:
        for key, value in row.items():
            if key.startswith("_"):
                continue
            if isinstance(value, bool):
                inferred = "boolean"
            elif isinstance(value, (int, float)):
                inferred = "number"
            elif isinstance(value, (dict, list)):
                inferred = "object"
            else:
                inferred = "string"
            if key not in types or types[key] == "string":
                types[key] = inferred
    return [{"name": name, "type": kind} for name, kind in types.items()]


# ----------------------------------------------------------------- dispatch


def list_all_sources() -> list[dict[str, Any]]:
    """Registered sources plus one implicit source per internal dataset."""
    implicit = [
        {
            "id": f"dataset:{entry['name']}",
            "name": entry["name"],
            "kind": "dataset",
            "config": {"dataset": entry["name"]},
            "rows": entry["rows"],
            "last_written_at": entry["last_written_at"],
        }
        for entry in get_dataset_store().list_datasets()
    ]
    return implicit + get_source_registry().list()


def resolve_source(source_id: str) -> dict[str, Any] | None:
    if source_id.startswith("dataset:"):
        name = source_id.split(":", 1)[1]
        return {
            "id": source_id,
            "name": name,
            "kind": "dataset",
            "config": {"dataset": name},
        }
    return get_source_registry().get(source_id)


def get_adapter(
    source: dict[str, Any],
) -> DatasetAdapter | SQLAdapter | FileAdapter:
    kind = source["kind"]
    config = source["config"]
    if kind == "dataset":
        return DatasetAdapter(config["dataset"])
    if kind == "sql":
        return SQLAdapter(
            config["credential"],
            table=config.get("table"),
            sql=config.get("sql"),
        )
    if kind == "file":
        return FileAdapter(
            config["file_id"], config.get("format", "xlsx"), config.get("sheet")
        )
    raise ValueError(f"Unknown source kind '{kind}'")


def list_credential_tables(credential: str) -> list[str]:
    """Table names visible through a stored SQL credential (for the UI)."""
    from sqlalchemy import create_engine, inspect

    entry = get_credential_store().get(credential)
    if entry is None or not entry.config.get("connection_string"):
        raise LookupError(f"Credential '{credential}' not found or not a database")
    engine = create_engine(str(entry.config["connection_string"]), pool_pre_ping=True)
    try:
        return sorted(inspect(engine).get_table_names())
    finally:
        engine.dispose()
